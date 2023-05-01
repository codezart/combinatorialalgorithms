from __init__ import *
from sudoku import SudokuPuzzle
from backtracking import BacktrackingSudoku,BacktrackingSudokuOptimized, MinigridBacktrackingSudoku
from ant_colony_optimization import AntColonyOptimizationSudoku,ACOAlgorithmSudoku
from bat_algorithm import BatAlgorithmSudoku
from genetic_algorithm import GeneticAlgorithmSudoku
from simulated_annealing import SimulatedAnnealingSudoku
from dancing_links import DancingLinksSudoku

puzzle = [
	[9,0,3,0,   0,2,0,8,    4,0,14,0,    0,1,0,7],
	[0,0,11,0,  0,0,9,4,    12,15,0,0,   0,5,0,0],
	[13,10,0,6, 0,0,16,0,   0,11,0,0,    8,0,12,3],
	[0,0,16,8,  11,14,12,0, 0,1,10,6,    2,15,0,0],
	
	[0,0,0,11,   7,0,0,0,    0,0,0,15,   10,0,0,0],
	[16,0,0,7,   0,9,3,0,    0,6,12,0,   15,0,0,13],
	[0,2,15,9,   0,8,0,0,    0,0,11,0,   6,7,1,0],
	[10,3,0,0,   0,0,0,11,   7,0,0,0,    0,0,5,16],
	
	[15,11,0,0,   0,0,0,9,  6,0,0,0,   0,0,8,12],
	[0,14,12,13,  0,1,0,0,  0,0,5,0,   11,3,15,0],
	[2,0,0,16,    0,4,7,0,  0,8,13,0,  14,0,0,1],
	[0,0,0,1,     3,0,0,0,  0,0,0,12,  4,0,0,0],
	
	[0,0,9,15,   12,10,1,0,  0,13,7,5,  3,2,0,0],
	[3,6,0,5,    0,0,2,0,    0,12,0,0,  1,0,9,14],
	[0,0,2,0,    0,0,14,13,  3,9,0,0,   0,0,0,0],
	[0,0,0,0,    0,0,0,0,    0,0,0,0,   0,0,0,0]
]
# puzzle = [
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 3, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 11, 0, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 4, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 6, 0, 0, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 16, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 12, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 6, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 7],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 11, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 9, 0, 0, 0, 0, 0, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 16, 0, 0, 0, 0, 0],
#     [0, 0, 8, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 6],
#     [0, 0, 0, 0, 0, 0, 0, 0, 10, 0, 0, 0, 0, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 8, 0, 0, 0],
#     [0, 0, 0, 0, 0, 0, 0, 0, 0, 4, 0, 0, 0, 0, 0, 0]
# ]

# puzzle_check =[
# 	[7, 9, 3, 6, 8, 4, 5, 1, 2], 
# 	[4, 8, 6, 5, 1, 2, 9, 3, 7], 
# 	[1, 2, 5, 9, 7, 3, 8, 4, 6], 

# 	[9, 3, 2, 7, 5, 1, 6, 8, 4], 
# 	[5, 7, 8, 2, 4, 6, 3, 9, 1], 
# 	[6, 4, 1, 3, 9, 8, 7, 2, 5], 

# 	[3, 1, 9, 4, 6, 5, 2, 7, 8], 
# 	[8, 5, 7, 1, 2, 9, 4, 6, 3], 
# 	[2, 6, 4, 8, 3, 7, 1, 5, 9]] 
# puzzle_check = [
# 	[1, 6, 3, 7, 2, 5, 9, 4, 8], 
# 	[5, 8, 4, 9, 5, 3, 2, 2, 1], 
# 	[7, 2, 9, 1, 8, 4, 3, 6, 5], 
# 	[9, 5, 6, 5, 7, 8, 1, 1, 2], 
# 	[4, 7, 1, 9, 6, 4, 2, 8, 9], 
# 	[6, 5, 2, 9, 1, 7, 5, 3, 4], 
# 	[4, 4, 8, 2, 8, 5, 7, 5, 6], 
# 	[6, 3, 7, 5, 4, 9, 8, 2, 2], 
# 	[2, 1, 5, 6, 8, 7, 3, 9, 3]]
puzzle_check = [[8, 6, 4, 3, 7, 1, 2, 5, 9], [3, 2, 5, 8, 4, 9, 7, 6, 1], [9, 7, 1, 2, 6, 5, 8, 4, 3], [4, 3, 6, 1, 9, 2, 5, 8, 7], [1, 9, 8, 6, 5, 7, 4, 3, 2], [2, 5, 7, 4, 8, 3, 9, 1, 6], [6, 8, 9, 7, 3, 4, 1, 2, 5], [7, 1, 3, 5, 2, 8, 6, 9, 4], [5, 4, 2, 9, 1, 6, 3, 7, 8]]
empty_puzzle = [
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0],
	[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
]

# Choose the desired algorithm by creating an instance of the corresponding class
size = 16  # or 25 for 25x25 puzzles

def check_sudoku(puzzle):
	sp = SudokuPuzzle(puzzle,size)
	for i in range(0,size):
		for j in range(0,size):
			if not sp.is_valid(i,j,sp.puzzle[i][j]):
				print(i,j, sp.puzzle[i][j])
				print("not valid")
				return	
	print("valid solution")





# BACKTRACKING
def backtracking(puzzle,size):
	logging.info("BACKTRACKING ALGORITHM")
	logging.info("MinigridBacktrackingSudoku Version")

	print("MinigridBacktrackingSudoku")
	solver = MinigridBacktrackingSudoku(puzzle, size)

	logging.info(f"size: {size}")
	logging.info(f"Initial puzzle: \n{puzzle}")

	start_time = time.process_time()
	
	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time

	logging.info("Final puzzle solved:")
	logging.info(solution)
	logging.info(f"Number of tries: {solver.numberoftries}")
	logging.info(f"Total Execution Time (s): {execution_time}")

	print(solution, solver.numberoftries)
	return {
		"puzzle": puzzle, 
		"solution": solution, 
		"numberoftries": solver.numberoftries, 
		"numof_is_valid_checks": solver.numof_is_valid_checks}

def ant_colony_optimization(puzzle,size):
	# num_ants=50
	# generations=500
	# alpha=0.9
	# beta=0.9
	# evaporation_rate=0.1
	num_ants=100
	generations=500
	alpha=0.8
	beta=1.2
	evaporation_rate=0.3
	print("ACOAlgorithmSudoku")

	logging.info(f"ANT COLONY OPTIMIZATION")
	logging.info(f"ACOAlgorithmSudoku VERSION constraint prop ")
	logging.info(f"num_ants: {num_ants}")
	logging.info(f"generations: {generations}")
	logging.info(f"alpha: {alpha}")
	logging.info(f"beta: {beta}")
	logging.info(f"evaporation_rate: {evaporation_rate}")
	logging.info(f"Initial puzzle: \n{puzzle}")

	solver = ACOAlgorithmSudoku(puzzle, size, num_ants, generations,alpha,beta ,evaporation_rate)

	start_time = time.process_time()

	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time
	
	logging.info("Final puzzle solved:")
	logging.info(solution)
	logging.info(f"Number of tries: {solver.numberoftries}")
	logging.info(f"Total Execution Time (s): {execution_time}")

	print(solution)
	return {
		"puzzle": puzzle,
		"solution":solution,
		"num_ants":num_ants,
		"alpha":alpha,
		"beta": beta,
		"evaporation_rate": evaporation_rate,
		"numberoftries": solver.numberoftries,
		"execution_time": execution_time,
		"numof_is_valid_checks": solver.numof_is_valid_checks}


def bat_algorithm(puzzle,size):
	# size = 16  # or 25 for 25x25 puzzles
	num_bats = 100
	generations = 100000
	alpha = 0.9
	gamma = 0.9
	print("BatAlgorithmSudoku")
	solver = BatAlgorithmSudoku(puzzle, size, num_bats, generations, alpha, gamma)
	start_time = time.process_time()

	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time
	print(solution)
	return {
		"puzzle": puzzle,
		"solution":solution,
		"generations":generations,
		"alpha":alpha,
		"gamma": gamma,
		"numberoftries": solver.numberoftries,
		"execution_time": execution_time,
		"numof_is_valid_checks": solver.numof_is_valid_checks}

def genetic_algorithm(puzzle,size):
	# size = 16  # or 25 for 25x25 puzzles
	# population_size = 1000
	# generations = 100000
	# mutation_rate = 0.1
	# crossover_rate = 0.2
	# 17 Givens
	population_size= 10000
	generations = 100000
	mutation_rate = 0.5
	crossover_rate = 0.7
	logging.info("GeneticAlgorithmSudokuElitist")
	logging.info("Elitist version")
	logging.info(f"population_size: {population_size}")
	logging.info(f"generations: {generations}")
	logging.info(f"mutation_rate: {mutation_rate}")
	logging.info(f"size: {size}")
	logging.info(f"Initial puzzle: \n{puzzle}")
	print("GeneticAlgorithmSudokuElitist")
	print(f"population_size: {population_size}")
	print(f"generations: {generations}")
	print(f"mutation_rate: {mutation_rate}")
	print(f"size: {size}")
	print(f"Initial puzzle: \n{puzzle}")

	solver = GeneticAlgorithmSudoku(puzzle, size, population_size, generations, mutation_rate, crossover_rate)
	start_time = time.process_time()

	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time
	if solution:
		logging.info(solution)
	else:
		logging.info("No solution found")
	return {
		"puzzle": puzzle,
		"solution":solution,
		"generations":generations,
		"population_size":population_size,
		"mutation_rate": mutation_rate,
		"numberoftries": solver.numberoftries,
		"solutions_generation": solver.solutions_generation,
		"execution_time": execution_time,
		"numof_is_valid_checks": solver.numof_is_valid_checks}

def simulated_annealing(puzzle, size):
	temperature=1000
	cooling_rate=0.999
	steps=1000000
	# size=16
	print("SimulatedAnnealingSudoku")
	logging.info("SimulatedAnnealingSudoku")
	logging.info("Generic version")
	logging.info(f"temperature: {temperature}")
	logging.info(f"cooling_rate: {cooling_rate}")
	logging.info(f"size: {size}")
	logging.info(f"Initial puzzle: \n{puzzle}")
	solver = SimulatedAnnealingSudoku(puzzle,size, temperature,cooling_rate,steps)
	start_time = time.process_time()

	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time
	logging.info("Final puzzle solved:")
	logging.info(solution)
	logging.info(f"Number of tries: {solver.numberoftries}")
	return {
		"puzzle": puzzle,
		"solution":solution,
		"temperature":temperature,
		"cooling_rate":cooling_rate,
		"steps": steps,
		"numberoftries": solver.numberoftries,
		"execution_time": execution_time,
		"numof_is_valid_checks": solver.numof_is_valid_checks}


def dancing_links(puzzle, size):
	logging.info("DancingLinksSudoku")
	logging.info("Generic version")
	logging.info(f"size: {size}")
	logging.info(f"Initial puzzle: \n{puzzle}")
	solver = DancingLinksSudoku(puzzle,size)
	start_time = time.process_time()

	solution = solver.solve()
	
	end_time = time.process_time()
	execution_time = end_time - start_time
	logging.info("Final puzzle solved:")
	logging.info(solution)
	return {
		"puzzle": puzzle,
		"solution":solution,
		"execution_time": execution_time,
	}






			
def string_to_2d_list(string, rows, cols):
	if rows * cols != len(string):
		raise ValueError("Size of the string does not match the given size of 2D list")
	lst = [[0] * cols for _ in range(rows)]
	for i in range(rows):
		for j in range(cols):
			lst[i][j] = int(string[i * cols + j])
	return lst

def get_datasetcsv(algorithm, size=9):
	logging.info('#############################')
	logging.info('RUNNING SUDOKU CSV 1 MILLION 9X9 PUZZLES')
	logging.info('#############################')

	import csv
	filename = "./datasets/sudoku.csv"   # replace with your file name
	column_name = "quizzes"    # replace with your column name
	# Open the file in read mode and create a CSV reader object
	size = 9
	results = {}
	with open(filename, "r") as file:
		reader = csv.DictReader(file)
		# Loop through the rows and print the values in the specified column	
		for index,row in enumerate(reader):
			puzzle = string_to_2d_list(row[column_name], size,size)
			results[index] = algorithm(puzzle=puzzle, size=size)
			logging.info('')
			logging.info('')
			if index == 100:
				logging.info("Final")
				logging.info(results)
				
				graphs_stats(results,algorithm)
				return results

def get_17givens_dataset(algorithm, size=9):
	logging.info('#############################')
	logging.info('RUNNING 17 GIVENS 9X9 PUZZLES')
	logging.info('#############################')
	results = {}
	filename = "./datasets/sudoku17.txt"
	with open(filename, "r") as file:
		for index,line in enumerate(file):
			puzzle = string_to_2d_list(line.strip(), size, size)
			results[index] = algorithm(puzzle=puzzle, size=size)
			logging.info('')
			logging.info('')
			if index == 50:
				logging.info("Final")
				logging.info(results)

				return results

def increment_version(folder_path,algorithm):
	# List all files in the folder
	files = os.listdir(folder_path)

	# Filter files that match the desired pattern
	pattern = re.compile(r'{}_v(\d+).log'.format(algorithm))
	matching_files = [f for f in files if pattern.match(f)]

	if matching_files:
		# Extract the version numbers and find the maximum
		versions = [int(pattern.match(f).group(1)) for f in matching_files]
		max_version = max(versions)

		# Increment the version number and create a new filename
		new_filename = f"{algorithm}_v{max_version + 1}.log"
		return new_filename
	else:
		return f"{algorithm}_v1.log"

def graphs_stats(results,algorithm):
	return
	# create a directed graph
	G = nx.DiGraph()

	# add nodes to the graph
	for key in results:
		G.add_node(key)

	# add edges to the graph
	for key, value in results.items():
		for child in value.get("children", []):
			G.add_edge(key, child)

	# draw the graph
	nx.draw(G, with_labels=True)
	# save the graph
	plt.savefig("graph.png")


def create_excel(results, algorithm):
	import openpyxl

	# Create a new workbook or load an existing one
	try:
		wb = openpyxl.load_workbook('my_file.xlsx')
	except FileNotFoundError:
		wb = openpyxl.Workbook()
		
	# Select the active worksheet
	ws = wb.active

	# Example dictionary with row keys and column-value dictionaries
	my_dict = {
		1: {'A': 'apple', 'B': 'banana', 'C': 'cherry'},
		2: {'A': 'apricot', 'B': 'blueberry', 'C': 'cranberry'}
	}

	# Loop over the dictionary and add rows and columns to the worksheet
	for row_key, col_dict in my_dict.items():
		for col_key, value in col_dict.items():
			ws[f'{col_key}{row_key}'] = value

	# Save the workbook
	wb.save('my_file.xlsx')




if __name__ == "__main__":
	algorithms = [
		"backtracking",
		"aco",
		"dl",
		"ga",
		"bat",
		"sa",
	]
	datasets = ["17givens","kaggle"]
	parser = argparse.ArgumentParser()
	parser.add_argument("algorithm")
	parser.add_argument("dataset")
	args = parser.parse_args()

	if not args.algorithm in algorithms:
		print(f"Error: Algorithm Arg (first arg) input should be one of these: {algorithms}")
		exit(1)

	if not args.dataset in datasets:
		print(f"Error: datasets Arg (second arg) input should be one of these: {datasets}")
		exit(1)

	# create log file
	folder_path = "./logs"  
	new_filename = increment_version(folder_path, args.algorithm)
	logging.basicConfig(filename=folder_path+"/"+new_filename, filemode='w', level="INFO", format='%(asctime)s - %(levelname)s - %(message)s')
	print(new_filename)
	size=9

	# 17 givens dataset
	if args.dataset == "17givens":
		if args.algorithm == "backtracking":
			get_17givens_dataset(backtracking)
		elif args.algorithm == "aco":
			get_17givens_dataset(ant_colony_optimization)

		elif args.algorithm == "sa":
			get_17givens_dataset(simulated_annealing)

		elif args.algorithm == "ga":
			get_17givens_dataset(genetic_algorithm)
		
		elif args.algorithm == "dl":
			get_17givens_dataset(dancing_links)

		else:
			get_17givens_dataset(bat_algorithm)

	# kaggle dataset
	else:
		if args.algorithm == "backtracking":
			get_datasetcsv(backtracking)
		elif args.algorithm == "aco":
			get_datasetcsv(ant_colony_optimization)

		elif args.algorithm == "sa":
			get_datasetcsv(simulated_annealing)

		elif args.algorithm == "ga":
			get_datasetcsv(genetic_algorithm)
		elif args.algorithm == "dl":
			get_datasetcsv(dancing_links)

		else:
			get_datasetcsv(bat_algorithm)

	# single puzzle
	# backtracking(puzzle=puzzle, size=size)
	# ant_colony_optimization(puzzle=puzzle,size=size)
	# simulated_annealing(puzzle=puzzle, size=size)
	# genetic_algorithm(puzzle=puzzle,size=size)	
	# bat_algorithm(puzzle=puzzle,size=size)
	# sp = SudokuPuzzle(puzzle_check,9)
	# total_errors = 0
	# for row in range(9):
	# 	for col in range(9):
	# 		if not sp.is_valid(row, col, sp.puzzle[row][col]):
	# 			total_errors += 1

	# print(total_errors)